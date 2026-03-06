from http.client import responses
from django.db.models import Q
from django.contrib.auth.models import Group
from django.utils import timezone
import csv
import io
from django.http import HttpResponse, HttpResponseBadRequest

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from django.utils.timezone import now as timezone_now
from reportlab.lib import colors
import logging
import io
from django.http import HttpResponse, HttpResponseServerError
from django.db.models import Count
from django.db.models.functions import Lower, Trim
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from django.db.models.functions import Lower, Trim

from reportlab.pdfgen import canvas
from django.shortcuts import render, redirect
from django.urls import reverse_lazy
from django.views.generic import TemplateView, View, CreateView, UpdateView, DeleteView
from django.contrib.auth import authenticate, login
from django.contrib.auth.mixins import LoginRequiredMixin
from .forms import UserRegisterForm, InventoryItemForm
from .models import InventoryItem, Category, TransactionLog, Status,SpecificItem
from inventory_management.settings import LOW_QUANTITY
from django.contrib import messages
from .forms import InventoryItemSearchForm,InventoryFullItemSearchForm
from django.shortcuts import get_object_or_404
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from django.views.generic import DetailView
from .forms import SpecificItemStatusForm
from django.http import JsonResponse

from django.db.models import Count
from django.db.models import Sum
from django.http import HttpResponseRedirect
from django.contrib.auth.mixins import UserPassesTestMixin
from django.urls import reverse
from django.views.generic import ListView
from django.db.models import Count, Q
from django.shortcuts import render
from django.http import HttpResponse, HttpResponseBadRequest
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView
from django.db.models import Q
import openpyxl
import io
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Lasvegas, Monaco, Suzuka, Abudhabi
from .models import SpecificItem, Room  # Import Room model
from datetime import datetime
from django import forms
import imgkit

class Index(TemplateView):

	template_name = 'inventory/index.html'
from django.db.models import Count, Q
from django.shortcuts import render
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from .models import InventoryItem, SpecificItem, Category
from .forms import InventoryItemSearchForm

from django.db.models import Count, Q
class Dashboard(LoginRequiredMixin, View):
    def get(self, request):
        inventory_items = InventoryItem.objects.all()
        categories = Category.objects.all()

        # Count all inventory items
        total_items = InventoryItem.objects.count()

        # Retrieve sorting parameters
        sort_by = request.GET.get('sort_by', 'date_modified')
        order = request.GET.get('order', 'desc')  # Default to descending order

        # Initialize the search form
        form = InventoryItemSearchForm(request.GET)
        query = form.cleaned_data.get('query', '') if form.is_valid() else None

        # Validate sorting fields
        valid_sort_fields = [
            'name', 'user', 'date_created', 'quantity', 'status', 'label',
            'category', 'supplier', 'room', 'local', 'date_modified', 'id', 'model',
        ]
        if sort_by not in valid_sort_fields:
            sort_by = 'date_modified'

        # Determine sorting order
        sort_field = f"-{sort_by}" if order == 'desc' else sort_by

        # Fetch inventory items based on user role
        if request.user.is_staff:
            items = InventoryItem.objects.all()
        else:
            items = InventoryItem.objects.filter(user=request.user)

        # Apply search query if present
        highlighted_item_ids = []
        if query:
            items = items.filter(
                Q(name__icontains=query) | Q(label__icontains=query) | Q(groups__name__icontains=query) | Q(model__icontains=query)
            )
            highlighted_item_ids = items.values_list('id', flat=True)

        # Apply sorting
        items = items.order_by(sort_field)

        # Identify low inventory items
        low_inventory = items.filter(quantity__lte=5)  # Adjust threshold if needed
        low_inventory_count = low_inventory.count()
        low_inventory_ids = low_inventory.values_list('id', flat=True)

        # Show a warning if there are low inventory items
        if low_inventory_count > 0:
            messages.error(request, f" {low_inventory_count} ")

        selected_category = request.GET.get('category')
        if selected_category:
            items = items.filter(category_id=selected_category)

        # Count totals for different statuses
        totals = SpecificItem.objects.aggregate(
            total_broken=Count('id', filter=Q(status='broken')),
            total_in_use=Count('id', filter=Q(status='inUse')),
            total_backup=Count('id', filter=Q(status='backup/Stock')),
        )

        # Count `inUse` items per inventory item
        in_use_counts = SpecificItem.objects.filter(status="inUse") \
            .values('inventory_item_id') \
            .annotate(count=Count('id'))
        in_use_counts_dict = {item['inventory_item_id']: item['count'] for item in in_use_counts}

        # Count backup stock per inventory item
        backup_counts = SpecificItem.objects.filter(status='backup/Stock') \
            .values('inventory_item_id') \
            .annotate(count=Count('id'))
        backup_counts_dict = {item['inventory_item_id']: item['count'] for item in backup_counts}
        # Count broken items per inventory item
        broken_counts = SpecificItem.objects.filter(status="broken") \
            .values('inventory_item_id') \
            .annotate(count=Count('id'))
        broken_counts_dict = {item['inventory_item_id']: item['count'] for item in broken_counts}

        return render(request, 'inventory/dashboard.html', {
            'items': items,
            'low_inventory_ids': low_inventory_ids,
            'form': form,
            'categories': categories,
            'highlighted_item_ids': highlighted_item_ids,
            'total_broken': totals.get('total_broken', 0),
            'total_in_use': totals.get('total_in_use', 0),
            'total_backup': totals.get('total_backup', 0),
            'backup_counts': backup_counts_dict,  # Send backup counts to template
            'in_use_counts': in_use_counts_dict,  # Send in-use counts to template
            'broken_counts': broken_counts_dict,  # Send broken counts to template
            'total_items': total_items,  # Send total items count to template
        })


import requests
from django.conf import settings
from django.contrib import messages
from django.shortcuts import render, redirect
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin
from .models import TransactionLog  # Ensure this import is correct


class TransactionLogsView(LoginRequiredMixin, View):
    TELEGRAM_BOT_TOKEN = "7915072048:AAE1ERblYSKRozSKzBucuPkHiyC4WU-_yis"
    TELEGRAM_CHAT_ID = "6043414970"  # Extracted from the JSON response

    def send_telegram_alert(self, message):
        url = f"https://api.telegram.org/bot{self.TELEGRAM_BOT_TOKEN}/sendMessage"
        payload = {"chat_id": self.TELEGRAM_CHAT_ID, "text": message}
        response = requests.post(url, json=payload)

        # Optional: Log any errors
        if response.status_code != 200:
            print(f"Error sending Telegram message: {response.text}")

    def get(self, request):
        logs = TransactionLog.objects.all().order_by('-timestamp')
        return render(request, 'inventory/transaction_logs.html', {'logs': logs})

    def post(self, request):
        # Clear all logs
        TransactionLog.objects.all().delete()
        messages.success(request, "All transaction logs have been cleared.")

        # Send Telegram alert
        self.send_telegram_alert("🚨 Transaction logs have been cleared!")

        return redirect('transaction-logs')


# Replace 'transaction-logs' with your URL name




from .models import SpecificItem  # Replace with your actual model
from reportlab.lib.pagesizes import letter

import logging
logger = logging.getLogger(__name__)
from django.http import HttpResponse, HttpResponseServerError
from django.db.models import Count
from django.db.models import Q
from django.http import HttpResponseBadRequest
from django.views.generic import ListView
from django.contrib.auth.mixins import LoginRequiredMixin
from .models import SpecificItem, Room

class SearchView(LoginRequiredMixin, ListView):
    model = SpecificItem
    template_name = 'inventory/search.html'
    context_object_name = 'specific_items'
    paginate_by = 10

    def get_queryset(self):
        """Filters and sorts the queryset based on request parameters."""
        room_name = self.request.GET.get('room', '')  # Room filter
        search_query = self.request.GET.get('search', '')  # Search query
        sort_by = self.request.GET.get('sort_by', 'label')  # Default sort by label
        order = self.request.GET.get('order', 'desc')  # Default order desc
        queryset = SpecificItem.objects.all()

        # Apply search filters
        if search_query:
            queryset = queryset.filter(

                Q(label__icontains=search_query) |
                Q(added_by__username__icontains=search_query) |
                Q(status__icontains=search_query) |
                Q(groups__name__icontains=search_query)
            )

        # Apply room filter if selected
        if room_name:
            queryset = queryset.filter(room__name__icontains=room_name)

        # Apply sorting
        valid_sort_fields = ['room', 'status', 'date_added', 'label']
        if sort_by in valid_sort_fields:
            sort_field = f"-{sort_by}" if order == 'desc' else sort_by
            queryset = queryset.order_by(sort_field)

        return queryset

    def get_context_data(self, **kwargs):
        """Pass additional context to the template, including item counts per room."""
        context = super().get_context_data(**kwargs)
        context['rooms'] = Room.objects.all()  # Pass all rooms
        context['selected_room'] = self.request.GET.get('room', '')  # Keep selected room

        # Count each item type in the selected room
        room_name = self.request.GET.get('room', None)
        if room_name:
            room_item_counts = (
                SpecificItem.objects
                .filter(room__name=room_name)
                .values("inventory_item__name")  # Group by item name
                .annotate(total=Count("id"))  # Count occurrences
                .order_by("-total")  # Order by highest count
            )
        else:
            room_item_counts = []

        context['room_item_counts'] = room_item_counts  # Add to context
        return context

    def get(self, request, *args, **kwargs):
        """Handles the GET request and checks for export formats."""
        format = request.GET.get('format', None)  # Export format (Excel or PDF)
        self.object_list = self.get_queryset()  # Get filtered items

        if format == 'excel':
            return self.export_to_excel(self.object_list)
        elif format == 'pdf':
            return self.export_to_pdf(self.object_list)
        elif format:
            return HttpResponseBadRequest("Invalid export format specified.")

        return self.render_to_response(self.get_context_data())


    import logging
    import openpyxl
    from django.http import HttpResponse, HttpResponseServerError
    from django.db.models import Count
    from django.db.models.functions import Lower, Trim

    logger = logging.getLogger(__name__)  # Ensure logger is defined

    def export_to_excel(self, queryset):
        """Exports filtered queryset to an Excel file with item counts in a separate table at the bottom."""
        try:
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="specific_items.xlsx"'

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = 'Inventory Data'

            # Headers for main inventory table
            headers = ['Inventory Item', 'Label', 'Status', 'Date Added', 'Groups', 'Room']
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num).value = header

            # Writing main inventory data
            row_num = 2
            for item in queryset:
                sheet.cell(row=row_num, column=1).value = item.inventory_item.name
                sheet.cell(row=row_num, column=2).value = item.label
                sheet.cell(row=row_num, column=3).value = item.status
                sheet.cell(row=row_num, column=4).value = item.date_added.strftime("%Y-%m-%d")
                sheet.cell(row=row_num, column=5).value = ', '.join(
                    [group.name for group in item.groups.all()]) or 'No Group'
                sheet.cell(row=row_num, column=6).value = item.room.name if item.room else 'No Room'
                row_num += 1

            # Leave a blank row
            row_num += 2

            # Headers for summary table
            sheet.cell(row=row_num, column=1).value = "Inventory Item"
            sheet.cell(row=row_num, column=2).value = "Total Count"

            # Count occurrences of each inventory item with normalization
            item_counts = (
                queryset
                .annotate(item_name_clean=Trim(Lower('inventory_item__name')))  # Normalize names
                .values('item_name_clean')  # Use the cleaned name for grouping
                .annotate(total=Count('id'))
            )

            # Writing summary table data
            row_num += 1
            for item in item_counts:
                sheet.cell(row=row_num, column=1).value = item[
                    'item_name_clean'].capitalize()  # Capitalize first letter
                sheet.cell(row=row_num, column=2).value = item['total']
                row_num += 1

            # Add grand total count
            sheet.cell(row=row_num + 1, column=1).value = "Grand Total"
            sheet.cell(row=row_num + 1, column=2).value = queryset.count()

            workbook.save(response)
            return response

        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            return HttpResponseServerError("Error exporting to Excel")

    logger = logging.getLogger(__name__)  # Ensure logger is defined

    def export_to_pdf(self, queryset):
        """Exports filtered queryset to a PDF file with item counts in a separate table at the bottom."""
        try:
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="specific_items.pdf"'

            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()

            # Main table headers
            data = [['Inventory Item', 'Label', 'Status', 'Date Added', 'Groups', 'Room']]

            # Add item data to the first table
            for item in queryset:
                data.append([
                    item.inventory_item.name,
                    item.label,
                    item.status,
                    item.date_added.strftime("%Y-%m-%d"),
                    ', '.join([group.name for group in item.groups.all()]) or 'No Group',
                    item.room.name if item.room else 'No Room'
                ])

            # Create main table
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))

            elements.append(table)
            elements.append(Spacer(1, 12))  # Add spacing between tables

            # Create summary table for total counts
            elements.append(Paragraph("Summary of Items", styles['Heading2']))

            summary_data = [['Item Name', 'Total Count']]  # Summary table header

            # Count occurrences of each inventory item with normalization
            item_counts = (
                queryset
                .annotate(item_name_clean=Trim(Lower('inventory_item__name')))  # Normalize names
                .values('item_name_clean')  # Use the cleaned name for grouping
                .annotate(total=Count('id'))
            )

            for item in item_counts:
                summary_data.append([item['item_name_clean'].capitalize(), item['total']])  # Capitalize first letter

            summary_table = Table(summary_data)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkgrey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))

            elements.append(summary_table)

            # Build PDF
            doc.build(elements)
            pdf = buffer.getvalue()
            buffer.close()
            response.write(pdf)
            return response

        except Exception as e:
            logger.error(f"PDF export failed: {e}")
            return HttpResponseServerError("Error exporting to PDF")

class SignUpView(View):
	def get(self, request):
		form = UserRegisterForm()
		return render(request, 'inventory/signup.html', {'form': form})

	def post(self, request):
		form = UserRegisterForm(request.POST)

		if form.is_valid():
			form.save()
			user = authenticate(
				username=form.cleaned_data['username'],
				password=form.cleaned_data['password1']
			)

			login(request, user)
			return redirect('index')


		return render(request, 'inventory/signup.html', {'form': form})


import requests
from django.conf import settings
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.utils import timezone
from django.views.generic import CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from .models import InventoryItem, SpecificItem, TransactionLog, Category
from .forms import InventoryItemForm


class TelegramNotificationMixin:
    """Mixin for sending Telegram alerts."""

    def send_telegram_alert(self, message):
        TELEGRAM_BOT_TOKEN = settings.TELEGRAM_BOT_TOKEN
        TELEGRAM_CHAT_ID = settings.TELEGRAM_CHAT_ID

        url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
        payload = {"chat_id": TELEGRAM_CHAT_ID, "text": message}

        try:
            response = requests.post(url, json=payload)
            response.raise_for_status()
        except requests.exceptions.RequestException as e:
            print(f"Error sending Telegram message: {e}")

class AddItem(LoginRequiredMixin, TelegramNotificationMixin, CreateView):
    model = InventoryItem
    form_class = InventoryItemForm
    template_name = 'inventory/item_form.html'
    success_url = reverse_lazy('dashboard')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categories'] = Category.objects.all()
        return context

    def form_valid(self, form):
        form.instance.user = self.request.user
        form.instance.name = form.instance.name.upper()
        form.instance.quantity = form.instance.quantity if form.instance.quantity and form.instance.quantity > 0 else 1
        form.user = self.request.user

        response = super().form_valid(form)

        # Get category, assigned group, supplier, and local details
        category_name = form.instance.category.name if form.instance.category else "No Category"
        assigned_group = form.instance.groups.name if form.instance.groups else "No Assigned Group"
        supplier_name = form.instance.supplier if form.instance.supplier else "No Supplier"
        local_name = form.instance.local if form.instance.local else "No Local"

        # Get Image URL
        image_url = self.request.build_absolute_uri(form.instance.image.url) if form.instance.image else None

        # Log transaction
        TransactionLog.objects.create(
            user=self.request.user,
            action='create',
            description=f"Created item: {form.instance.name} (Quantity: {form.instance.quantity}) - "
                        f"Category: {category_name}, Supplier: {supplier_name}, Local: {local_name}, "
                        f"Assigned to: {assigned_group}",
        )

        # Build Telegram message
        message = (
            f"✅ Item Created: {form.instance.name} (Quantity: {form.instance.quantity})\n"
            f"✅ Category: {category_name}\n"
            f"✅ Supplier: {supplier_name}\n"
            f"✅ Local: {local_name}\n"
            f"✅ Assigned to: {assigned_group}\n"
            f"✅ Created by: {self.request.user.username}"
        )

        # Send Telegram alert with image if available
        if image_url:
            message += f"\n✅ Image: [View Image]({image_url})"

        self.send_telegram_alert(message)

        return response


class EditItem(LoginRequiredMixin, TelegramNotificationMixin, UpdateView):
    model = InventoryItem
    form_class = InventoryItemForm
    template_name = 'inventory/item_form.html'
    success_url = reverse_lazy('dashboard')


    def get_success_url(self):
        # Redirect to item detail page after saving
        return reverse('item-detail', kwargs={'pk': self.object.pk})


    def form_valid(self, form):
        form.instance.user = self.request.user
        form.instance.date_modified = timezone.now()
        original_item = self.get_object()

        response = super().form_valid(form)

        if form.instance.quantity != original_item.quantity:
            old_quantity = original_item.quantity
            new_quantity = form.instance.quantity
            TransactionLog.objects.create(
                user=self.request.user,
                action='update',
                description=f"Updated quantity for {form.instance.name}: {old_quantity} → {new_quantity}",
            )

            self.send_telegram_alert(
                f"✅ Item Updated: {form.instance.name} - Quantity changed from {old_quantity} to {new_quantity} by {self.request.user.username}")

        return response

    def get_object(self, queryset=None):
        if self.request.user.is_staff:
            return get_object_or_404(InventoryItem, id=self.kwargs['pk'])
        return get_object_or_404(InventoryItem, id=self.kwargs['pk'], user=self.request.user)


class DeleteItem(LoginRequiredMixin, TelegramNotificationMixin, DeleteView):
    model = SpecificItem
    template_name = 'inventory/delete_specific_item_confirm.html'
    success_url = reverse_lazy('dashboard')
    context_object_name = 'specific_item'

    def get_object(self, queryset=None):
        if self.request.user.is_superuser:
            return get_object_or_404(SpecificItem, id=self.kwargs['pk'])
        return get_object_or_404(SpecificItem, id=self.kwargs['pk'], added_by=self.request.user)

    def delete(self, request, *args, **kwargs):
        obj = self.get_object()
        item_name = obj.name

        response = super().delete(request, *args, **kwargs)

        TransactionLog.objects.create(
            user=self.request.user,
            action='delete',
            description=f"Deleted item: {item_name}",
        )

        self.send_telegram_alert(f"🚨 Item Deleted: {item_name} by {self.request.user.username}")

        return response
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from .models import InventoryItem, SpecificItem

class ExportToExcelView(LoginRequiredMixin, View):
    def get(self, request):
        sort_by = request.GET.get('sort_by', 'name')
        query = request.GET.get('query', '')

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="inventory_data.xlsx"'

        # Create workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Inventory Data'

        # Excel headers
        headers = [
            'ID', 'Name', 'Label', 'Total Items', 'Total In Use',
            'Backup', 'Broken', 'Quantity', 'Date Created', 'Model'
        ]
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)

        # Fetch Inventory Items
        if request.user.is_staff:
            items = InventoryItem.objects.all()
        else:
            items = InventoryItem.objects.filter(user=request.user)

        if query:
            items = items.filter(name__icontains=query)

        valid_sort_fields = ['name', 'label', 'date_created', 'quantity', 'model']
        if sort_by not in valid_sort_fields:
            sort_by = 'name'
        items = items.order_by(sort_by)

        # Precompute counts for performance
        total_items_dict = {}
        total_in_use_dict = {}
        total_backup_dict = {}
        total_broken_dict = {}

        for item in items:
            # Safely get SpecificItems
            specific_qs = getattr(item, 'specific_items', None)
            if specific_qs is None:
                specific_qs = SpecificItem.objects.filter(inventory_item=item)

            total_items_dict[item.id] = specific_qs.count()
            total_in_use_dict[item.id] = specific_qs.filter(status__iexact='inUse').count()
            total_backup_dict[item.id] = specific_qs.filter(status__iexact='backup/Stock').count()
            total_broken_dict[item.id] = specific_qs.filter(status__iexact='Broken').count()

        # Write rows to Excel
        for row_num, item in enumerate(items, start=2):
            sheet.cell(row=row_num, column=1).value = item.id
            sheet.cell(row=row_num, column=2).value = item.name
            sheet.cell(row=row_num, column=3).value = getattr(item, 'label', '')
            sheet.cell(row=row_num, column=4).value = total_items_dict.get(item.id, 0)
            sheet.cell(row=row_num, column=5).value = total_in_use_dict.get(item.id, 0)
            sheet.cell(row=row_num, column=6).value = total_backup_dict.get(item.id, 0)
            sheet.cell(row=row_num, column=7).value = total_broken_dict.get(item.id, 0)
            sheet.cell(row=row_num, column=8).value = item.quantity
            sheet.cell(row=row_num, column=9).value = item.date_created.strftime('%Y-%m-%d') if item.date_created else ''
            sheet.cell(row=row_num, column=10).value = getattr(item, 'model', '')

        workbook.save(response)
        return response

class StatusChartView(View):
    def get(self, request):
        statuses = Status.objects.all()
        items = InventoryItem.objects.all()

        # Build labels and totals dynamically
        status_labels = [status.name for status in statuses]
        status_totals = [
            items.filter(status=status).count() for status in statuses
        ]

        return render(request, 'inventory/charts.html', {
            'status_labels': status_labels,
            'status_totals': status_totals,
        })



class StatusSpecificItemsView(View):
    def get(self, request):
        statuses = Status.objects.all()
        items = SpecificItem.objects.all()

        # Build labels and totals dynamically
        # Ensure data is correctly calculated
        status_labels = [status.name for status in statuses]
        status_totals = [items.filter(status=status).count() for status in statuses]

        # If SpecificItem does not have a direct room field but references InventoryItem's room


        return render(request, 'inventory/charts1.html', {
            'status_labels': status_labels,
            'status_totals': status_totals,

        })




from django.db import transaction

class ItemDetailView(LoginRequiredMixin, DetailView):
    model = InventoryItem
    template_name = 'inventory/item_detail.html'
    context_object_name = 'item'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        item = context['item']
        specific_items = SpecificItem.objects.filter(inventory_item=item)

        # Get search and sorting parameters
        query = self.request.GET.get('search', '')
        sort_by = self.request.GET.get('sort_by', 'date_added')
        order = self.request.GET.get('order', 'desc')

        # Apply sorting
        valid_sort_fields = ['label', 'date_added']
        if sort_by in valid_sort_fields:
            sort_field = f"-{sort_by}" if order == 'desc' else sort_by
            specific_items = specific_items.order_by(sort_field)

        # Apply search filter
        if query:
            specific_items = specific_items.filter(
                Q(label__icontains=query) |
                Q(added_by__username__icontains=query) |
                Q(status__icontains=query)
            )

        STATUS_MAPPING = dict(SpecificItem.STATUS_CHOICES)

        # Query to count items based on model and status
        model_status_counts = (
            specific_items.values('model', 'status')
            .annotate(count=Count('id'))
            .order_by('model', 'status')
        )

        model_data = {}
        total_model_counts = {}

        for entry in model_status_counts:
            model_name = entry["model"] or "Unknown Model"
            status_key = entry["status"] or "No Status"
            status_label = STATUS_MAPPING.get(status_key, status_key)

            count = entry["count"]

            if model_name not in model_data:
                model_data[model_name] = {}

            model_data[model_name][status_label] = count
            total_model_counts[model_name] = total_model_counts.get(model_name, 0) + count

        # Convert queryset to dictionary for easier template usage

        # Calculate statistics for the item

#------------------Inuse---------------------------------------#
        las_vegas_count = specific_items.filter(
    Q(room__name="S1 (LAS VEGAS)") ,status='inUse').count()
        abu_dhabi_count =specific_items.filter(
    Q(room__name="S2 (ABU DHABI)") ,status='inUse').count()
        monaco_count = specific_items.filter(
    Q(room__name="S3 (MONACO)") ,status='inUse').count()
        suzuka_count = specific_items.filter(
    Q(room__name="S4 (SUZUKA)") ,status='inUse').count()
        it_room_count = specific_items.filter(
    Q(room__name="IT-Office"), status='inUse').count()
        it_room_storage_count = specific_items.filter(
    Q(room__name="IT-Room-(Stockroom)"), status='inUse').count()
        c1_las_vegas_count = specific_items.filter(
    Q(room__name="C1 (LAS VEGAS)"), status='inUse').count()
        c2_abu_dhabi_count = specific_items.filter(
    Q(room__name="C2 (ABU DHABI)"), status='inUse').count()
        c3_suzuka_count = specific_items.filter(
    Q(room__name="C3 (SUZUKA)"), status='inUse').count()
        c4_monaco_count = specific_items.filter(
    Q(room__name="C4 (MONACO)"), status='inUse').count()
        storage_1_count = specific_items.filter(
    Q(room__name="Storage"), status='inUse').count()
        admin_hr_count = specific_items.filter(
    Q(room__name="Admin/Hr"), status='inUse').count()
        hallway_count = specific_items.filter(
    Q(room__name="Hallway"), status='inUse').count()
        commentator_pantry_count = specific_items.filter(
    Q(room__name="Commentator-Pantry"), status='inUse').count()
        ph_studio_count = specific_items.filter(
            Q(room__name="PH-STUDIO"), status='inUse').count()
        artist_dept_count = specific_items.filter(
            Q(room__name="3D-Artist-Dept"), status='inUse').count()
        seven_floor_count = specific_items.filter(
            Q(room__name="7TH-FLOOR"), status='inUse').count()
        plinko_count = specific_items.filter(
            Q(room__name="Plinko"), status='inUse').count()
        mini_pitx_count = specific_items.filter(
            Q(room__name="MINI-PITX-TRACK"), status='inUse').count()
        #----------------------readyForDeployment-----------------------#

        las_vegas_count_deployment = specific_items.filter(
    Q(room__name="S1 (LAS VEGAS)") , status='readyForDeployment'
).count()
        abu_dhabi_count_deployment = specific_items.filter(
    Q(room__name="S2 (ABU DHABI)") , status='readyForDeployment'
).count()
        monaco_count_deployment = specific_items.filter(
    Q(room__name="S3 (MONACO)") , status='readyForDeployment'
).count()
        suzuka_count_deployment = specific_items.filter(
    Q(room__name="S4 (SUZUKA)") , status='readyForDeployment'
).count()
        c1_las_vegas_count_deployment = specific_items.filter(
            Q(room__name="C1 (LAS VEGAS)"), status='readyForDeployment'
).count()
        c2_abu_dhabi_count_deployment = specific_items.filter(
            Q(room__name="C2 (ABU DHABI)"), status='readyForDeployment'
).count()
        c3_suzuka_count_deployment = specific_items.filter(
            Q(room__name="C3 (SUZUKA)"), status='readyForDeployment'
).count()
        c4_monaco_count_deployment = specific_items.filter(
             Q(room__name="C4 (MONACO)"), status='readyForDeployment'
).count()

        it_room_count_deployment = specific_items.filter(
            Q(room__name="IT-Office"), status='readyForDeployment'
).count()
        it_room_storage_count_deployment = specific_items.filter(
            Q(room__name="IT-Room-(Stockroom)"), status='readyForDeployment'
        ).count()
        storage_1_count_deployment = specific_items.filter(
            Q(room__name="Storage"), status='readyForDeployment'
).count()
        admin_hr_count_deployment = specific_items.filter(
            Q(room__name="Admin/Hr"), status='readyForDeployment'
).count()
        hallway_count_deployment = specific_items.filter(
            Q(room__name="Hallway"), status='readyForDeployment'
).count()
        commentator_pantry_count_deployment = specific_items.filter(
            Q(room__name="Commentator-Pantry"), status='readyForDeployment'
).count()
        ph_studio_count_deployment = specific_items.filter(
            Q(room__name="PH-STUDIO"), status='readyForDeployment'
).count()
        artist_dept_count_deployment = specific_items.filter(
            Q(room__name="3D-Artist-Dept"), status='readyForDeployment'
).count()
        seven_floor_count_deployment = specific_items.filter(
            Q(room__name="7TH-FlOOR"), status='readyForDeployment'
).count()
        plinko_count_deployment = specific_items.filter(
            Q(room__name="Plinko"), status='readyForDeployment'
        ).count()
        mini_pitx_count_deployment = specific_items.filter(
            Q(room__name="MINI-PITX-TRACK"), status='readyForDeployment'
).count()
        #-------------------backup/Stock------------------------------#

        las_vegas_count_backup = specific_items.filter(
            Q(room__name="S1 (LAS VEGAS)"), status='backup/Stock'
        ).count()

        abu_dhabi_count_backup = specific_items.filter(
            Q(room__name="S2 (ABU DHABI)"), status='backup/Stock'
        ).count()

        monaco_count_backup = specific_items.filter(
            Q(room__name="S3 (MONACO)"), status='backup/Stock'
        ).count()

        suzuka_count_backup = specific_items.filter(
            Q(room__name="S4 (SUZUKA)"), status='backup/Stock'
        ).count()
        c1_las_vegas_count_backup = specific_items.filter(
            Q(room__name="S1 (LAS VEGAS)"), status='backup/Stock'
        ).count()

        c2_abu_dhabi_count_backup = specific_items.filter(
            Q(room__name="S2 (ABU DHABI)"), status='backup/Stock'
        ).count()

        c4_monaco_count_backup = specific_items.filter(
            Q(room__name="S3 (MONACO)"), status='backup/Stock'
        ).count()

        c3_suzuka_count_backup = specific_items.filter(
            Q(room__name="S4 (SUZUKA)"), status='backup/Stock'
        ).count()

        it_room_count_backup = specific_items.filter(
            Q(room__name="IT-Office"), status='backup/Stock'
        ).count()
        it_room_storage_count_backup = specific_items.filter(
            Q(room__name="IT-Room-(Stockroom)"), status='backup/Stock'
        ).count()
        storage_1_count_backup = specific_items.filter(
            Q(room__name="Storage"), status='backup/Stock'
        ).count()
        admin_hr_count_backup = specific_items.filter(
            Q(room__name="Admin/Hr"), status='backup/Stock'
        ).count()
        hallway_count_backup = specific_items.filter(
            Q(room__name="Hallway"), status='backup/Stock'
        ).count()
        commentator_pantry_count_backup = specific_items.filter(
            Q(room__name="Commentator-Pantry"), status='backup/Stock'
        ).count()
        ph_studio_count_backup = specific_items.filter(
            Q(room__name="PH-STUDIO"), status='backup/Stock'
        ).count()
        artist_dept_count_backup = specific_items.filter(
            Q(room__name="3D-Artist-Dept"), status='backup/Stock'
        ).count()
        seven_floor_count_backup = specific_items.filter(
            Q(room__name="7TH-FlOOR"), status='backup/Stock'
        ).count()
        plinko_count_backup = specific_items.filter(
            Q(room__name="Plinko"), status='backup/Stock'
        ).count()
        mini_pitx_count_backup = specific_items.filter(
            Q(room__name="MINI-PITX-TRACK"), status='backup/Stock'
        ).count()


        las_vegas_ready_backup = las_vegas_count_deployment + las_vegas_count_backup
        abu_dhabi_ready_backup = abu_dhabi_count_deployment + abu_dhabi_count_backup
        suzuka_ready_backup =  suzuka_count_deployment +   suzuka_count_backup
        monaco_ready_backup =  monaco_count_deployment + monaco_count_backup
        c1_las_vegas_ready_backup = c1_las_vegas_count_deployment + c1_las_vegas_count_backup
        c2_abu_dhabi_ready_backup = c2_abu_dhabi_count_deployment + c2_abu_dhabi_count_backup
        c3_suzuka_ready_backup = c3_suzuka_count_deployment + c3_suzuka_count_backup
        c4_monaco_ready_backup = c4_monaco_count_deployment + c4_monaco_count_backup
        it_room_ready_backup = it_room_count_deployment + it_room_count_backup
        it_room_storage_ready_backup = it_room_storage_count_deployment + it_room_storage_count_backup
        storage_1_ready_backup = storage_1_count_deployment + storage_1_count_backup
        admin_hr_ready_backup = admin_hr_count_deployment + admin_hr_count_backup
        hallway_ready_backup = hallway_count_deployment + hallway_count_backup
        commentator_pantry_ready_backup = commentator_pantry_count_deployment + commentator_pantry_count_backup
        ph_studio_ready_backup = ph_studio_count_deployment + ph_studio_count_backup
        artist_dept_ready_backup = artist_dept_count_deployment + artist_dept_count_backup
        seven_floor_ready_backup = seven_floor_count_deployment + seven_floor_count_backup
        mini_pitx_ready_backup = mini_pitx_count_deployment + mini_pitx_count_backup
        plinko_ready_backup = plinko_count_deployment + plinko_count_backup






        #items computation--------------------------------------------------------------


        current_las_vegas =  int(item.requirement_las_vegas or 0) + int(item.propose_spare_las_vegas or 0)
        current_propose =  las_vegas_count + las_vegas_ready_backup
        las_vegas_total = max(0, current_las_vegas - current_propose)

        current_abu_dhabi = int(item.requirement_abu_dhabi or 0) + int(item.propose_spare_abu_dhabi or 0)
        current_propose_abu_dhabi = abu_dhabi_count + abu_dhabi_ready_backup
        abu_dhabi_total = max(0,current_abu_dhabi - current_propose_abu_dhabi)

        current_monaco = int(item.requirement_monaco or 0) + int(item.propose_spare_monaco or 0)
        current_propose_monaco =  monaco_count + monaco_ready_backup
        monaco_total = max(0,current_monaco - current_propose_monaco)

        current_suzuka = int(item.requirement_suzuka or 0) + int(item.propose_spare_suzuka or 0)
        current_propose_suzuka = suzuka_count + suzuka_ready_backup
        suzuka_total = max(0,current_suzuka - current_propose_suzuka)

        c1_current_las_vegas =  int(item.c1_requirement_las_vegas or 0) + int(item.c1_propose_spare_las_vegas or 0)
        c1_current_propose =  c1_las_vegas_count + c1_las_vegas_ready_backup
        c1_las_vegas_total = max(0, c1_current_las_vegas - c1_current_propose)

        c2_current_abu_dhabi = int(item.c2_requirement_abu_dhabi or 0) + int(item.c2_propose_spare_abu_dhabi or 0)
        c2_current_propose_abu_dhabi = c2_abu_dhabi_count + c2_abu_dhabi_ready_backup
        c2_abu_dhabi_total = max(0,c2_current_abu_dhabi - c2_current_propose_abu_dhabi)


        c4_current_monaco = int(item.c4_requirement_monaco or 0) + int(item.c4_propose_spare_monaco or 0)
        c4_current_propose_monaco =  c4_monaco_count + c4_monaco_ready_backup
        c4_monaco_total = max(0,c4_current_monaco - c4_current_propose_monaco)

        c3_current_suzuka = int(item.c3_requirement_suzuka or 0) + int(item.c3_propose_spare_suzuka or 0)
        c3_current_propose_suzuka = c3_suzuka_count + c3_suzuka_ready_backup
        c3_suzuka_total = max(0,c3_current_suzuka - c3_current_propose_suzuka)

        it_room_current = int(item.it_room_requirements or 0) + int(item.it_room_propose_spare or 0)
        it_room_current_propose_spare = it_room_count + it_room_ready_backup
        it_room_total = max(0,it_room_current - it_room_current_propose_spare)


        it_room_storage_current = int(item.it_room_storage_requirement or 0) + int(item.it_room_storage_propose_spare or 0)
        it_room_storage_current_propose_spare = it_room_storage_count + it_room_storage_ready_backup
        it_room_storage_total = max(0,it_room_storage_current - it_room_storage_current_propose_spare)


        storage_1_current = int(item.storage_1_requirement or 0) + int(item.storage_1_propose_spare or 0)
        storage_1_current_propose_spare = storage_1_count + storage_1_ready_backup
        storage_1_total = max(0,storage_1_current - storage_1_current_propose_spare)

        admin_hr_current = int(item.admin_hr_requirement or 0) + int(item.admin_hr_propose_spare or 0)
        admin_hr_current_propose_spare = admin_hr_count + admin_hr_ready_backup
        admin_hr_total = max(0,admin_hr_current - admin_hr_current_propose_spare)

        hallway_current = int(item.hallway_requirement or 0) + int(item.hallway_propose_spare or 0)
        hallway_current_propose_spare = hallway_count + hallway_ready_backup
        hallway_total = max(0,hallway_current - hallway_current_propose_spare)

        commentator_pantry_current = int(item.commentator_pantry_requirement or 0) + int(item.commentator_pantry_propose_spare or 0)
        commentator_pantry_current_propose_spare = commentator_pantry_count + commentator_pantry_ready_backup
        commentator_pantry_total = max(0,commentator_pantry_current - commentator_pantry_current_propose_spare)

        ph_studio_current = int(item.requirement_ph_studio or 0) + int(item.propose_spare_ph_studio or 0)
        ph_studio_current_propose_spare = ph_studio_count + ph_studio_ready_backup
        ph_studio_total = max(0, ph_studio_current - ph_studio_current_propose_spare)

        artist_dept_current = int(item.requirement_3d_dept or 0) + int(item.propose_spare_3d_dept or 0)
        artist_dept_current_propose_spare = artist_dept_count + artist_dept_ready_backup
        artist_dept_total = max(0, artist_dept_current - artist_dept_current_propose_spare)

        seven_floor_current = int(item.requirement_seven_floor or 0) + int(item.propose_spare_seven_floor or 0)
        seven_floor_current_propose_spare = seven_floor_count + seven_floor_ready_backup
        seven_floor_total = max(0, seven_floor_current - seven_floor_current_propose_spare)


        mini_pitx_current = int(item.requirement_mini_pitx or 0) + int(item.propose_spare_mini_pitx or 0)
        mini_pitx_current_propose_spare = mini_pitx_count + mini_pitx_ready_backup
        mini_pitx_total = max(0, mini_pitx_current - mini_pitx_current_propose_spare)

        plinko_current = int(item.requirement_mini_pitx or 0) + int(item.propose_spare_mini_pitx or 0)
        plinko_current_propose_spare = mini_pitx_count + mini_pitx_ready_backup
        plinko_total = max(0, mini_pitx_current - mini_pitx_current_propose_spare)





        las_vegas_requirement_propose = int(item.requirement_las_vegas or 0) + int(item.propose_spare_las_vegas or 0)
        abu_dhabi_requirement_propose = int(item.requirement_abu_dhabi or 0) + int(item.propose_spare_abu_dhabi or 0)
        monaco_requirement_propose = int(item.requirement_monaco or 0) + int(item.propose_spare_monaco or 0)
        suzuka_requirement_propose = int(item.requirement_suzuka or 0) + int(item.propose_spare_suzuka or 0)
        c1_las_vegas_requirement_propose = int(item.c1_requirement_las_vegas or 0) + int(item.c1_propose_spare_las_vegas or 0)
        c2_abu_dhabi_requirement_propose = int(item.c2_requirement_abu_dhabi or 0) + int(item.c2_propose_spare_abu_dhabi or 0)
        c4_monaco_requirement_propose = int(item.c4_requirement_monaco or 0) + int(item.c4_propose_spare_monaco or 0)
        c3_suzuka_requirement_propose = int(item.c3_requirement_suzuka or 0) + int(item.c3_propose_spare_suzuka or 0)
        it_room_requirement_propose = int(item.it_room_requirements or 0) + int(item.it_room_propose_spare or 0)
        it_room_storage_requirement_propose = int(item.it_room_storage_requirement or 0) + int(item.it_room_storage_propose_spare or 0)
        storage_1_requirement_propose = int(item.storage_1_requirement or 0) + int(item.storage_1_propose_spare or 0)
        admin_hr_requirement_propose = int(item.admin_hr_requirement or 0) + int(item.admin_hr_propose_spare or 0)
        hallway_requirement_propose = int(item.hallway_requirement or 0) + int(item.hallway_propose_spare or 0)
        commentator_pantry_requirement_propose = int(item.commentator_pantry_requirement or 0) + int(item.commentator_pantry_propose_spare or 0)
        ph_studio_requirement_propose = int(item.requirement_ph_studio or 0) + int(item.propose_spare_ph_studio or 0)
        artist_dept_requirement_propose = int(item.requirement_3d_dept or 0) + int(item.propose_spare_3d_dept or 0)
        seven_floor_requirement_propose = int(item.requirement_seven_floor or 0) + int(item.propose_spare_seven_floor or 0)
        mini_pitx_requirement_propose = int(item.requirement_mini_pitx or 0) + int(item.propose_spare_mini_pitx or 0)
        plinko_requirement_propose = int(item.requirement_plinko or 0) + int(item.propose_spare_plinko or 0)

        total_in_use = specific_items.filter(status='inUse').count()

        total_broken = specific_items.filter(status='broken').count()
        total_backup = specific_items.filter(status='backup/Stock').count()

        total_none = specific_items.filter(status='None').count()
        total_ready = specific_items.filter(status='readyForDeployment').count()







        #fixed
        total_spare = (las_vegas_count_deployment + abu_dhabi_count_deployment + monaco_count_deployment + suzuka_count_deployment + it_room_count_deployment + c1_las_vegas_count_deployment + c2_abu_dhabi_count_deployment + c3_suzuka_count_deployment + c4_monaco_count_deployment + storage_1_count_deployment + admin_hr_count_deployment + hallway_count_deployment + it_room_storage_count_deployment + commentator_pantry_count_deployment + ph_studio_count_deployment + artist_dept_count_deployment + seven_floor_count_deployment + mini_pitx_count_deployment + plinko_count_deployment)
        total_needed = las_vegas_requirement_propose + abu_dhabi_requirement_propose + monaco_requirement_propose + suzuka_requirement_propose + c1_las_vegas_requirement_propose + c2_abu_dhabi_requirement_propose + c4_monaco_requirement_propose + c3_suzuka_requirement_propose + it_room_requirement_propose + storage_1_requirement_propose + admin_hr_requirement_propose + hallway_requirement_propose + it_room_storage_requirement_propose + commentator_pantry_requirement_propose + ph_studio_requirement_propose + artist_dept_requirement_propose + seven_floor_requirement_propose + mini_pitx_requirement_propose + plinko_requirement_propose
        #fixed
        quantity_needed = las_vegas_total + abu_dhabi_total + suzuka_total + monaco_total + c1_las_vegas_total + c2_abu_dhabi_total + c3_suzuka_total + c4_monaco_total + it_room_total + storage_1_total + admin_hr_total + hallway_total + it_room_storage_total + commentator_pantry_total + ph_studio_total + artist_dept_total + seven_floor_total + mini_pitx_total + plinko_total



        overall_spare = (total_backup + total_ready)
        total_quantity = total_in_use + total_spare + total_backup + total_broken


        total_items = total_spare + total_backup
        total_needed_to_buy = total_items - total_needed

        if overall_spare >= quantity_needed:
            result = 0
        else:
            result = quantity_needed - overall_spare


        context.update({

            'result':result,
            'specific_items': specific_items,
            'model_data': model_data,  # Dictionary with model -> status counts
            'total_model_counts': total_model_counts,
            "STATUS_CHOICES": SpecificItem.STATUS_CHOICES,  # Pass choices to template
            'total_ready': total_ready,
            'total_in_use': total_in_use,
            'total_broken': total_broken,
            'total_backup': total_backup,
            'total_none': total_none,
            'total_items':total_items,
            'quantity_needed':quantity_needed,
            'total_quantity':total_quantity,
            #'need_to_order': need_to_order,
            'current_sort': sort_by,

            #room count
            'las_vegas_count': las_vegas_count,
            'abu_dhabi_count': abu_dhabi_count,
            'suzuka_count': suzuka_count,
            'monaco_count': monaco_count,
            'c1_las_vegas_count': c1_las_vegas_count,
            'c2_abu_dhabi_count': c2_abu_dhabi_count,
            'c3_suzuka_count': c3_suzuka_count,
            'c4_monaco_count': c4_monaco_count,
            'it_room_count': it_room_count,
            'storage_1_count':storage_1_count,
            'admin_hr_count':admin_hr_count,
            'hallway_count':hallway_count,
            'it_room_storage_count':it_room_storage_count,
            'commentator_pantry_count': commentator_pantry_count,
            'ph_studio_count': ph_studio_count,
            'artist_dept_count': artist_dept_count,
            'seven_floor_count': seven_floor_count,
            'mini_pitx_count': mini_pitx_count,
            'plinko_count': plinko_count,
            #'other_room_count': other_room_count,


            #deployment count
            'las_vegas_count_deployment': las_vegas_count_deployment,
            'abu_dhabi_count_deployment': abu_dhabi_count_deployment,
            'monaco_count_deployment': monaco_count_deployment,
            'suzuka_count_deployment': suzuka_count_deployment,
            'c1_las_vegas_count_deployment': c1_las_vegas_count_deployment,
            'c2_abu_dhabi_count_deployment': c2_abu_dhabi_count_deployment,
            'c3_suzuka_count_deployment': c3_suzuka_count_deployment,
            'c4_monaco_count_deployment': c4_monaco_count_deployment,
            'it_room_count_deployment': it_room_count_deployment,
            'storage_1_count_deployment': storage_1_count_deployment,
            'admin_hr_count_deployment':admin_hr_count_deployment,
            'hallway_count_deployment': hallway_count_deployment,
            'it_room_storage_count_deployment':it_room_storage_count_deployment,
            'commentator_pantry_count_deployment':commentator_pantry_count_deployment,
            'ph_studio_count_deployment':ph_studio_count_deployment,
            'artist_dept_count_deployment':artist_dept_count_deployment,
            'seven_floor_count_deployment':seven_floor_count_deployment,
            'mini_pitx_count_deployment': mini_pitx_count_deployment,
            'plinko_count_deployment':plinko_count_deployment,

            #backup count
            'las_vegas_count_backup':las_vegas_count_backup,
            'abu_dhabi_count_backup':abu_dhabi_count_backup,
            'suzuka_count_backup': suzuka_count_backup,
            'monaco_count_backup':monaco_count_backup,
            'c1_las_vegas_count_backup':c1_las_vegas_count_backup,
            'c2_abu_dhabi_count_backup':c2_abu_dhabi_count_backup,
            'c3_suzuka_count_backup':c3_suzuka_count_backup,
            'c4_monaco_count_backup':c4_monaco_count_backup,
            'it_room_count_backup':it_room_count_backup,
            'storage_1_count_backup':storage_1_count_backup,
            'admin_hr_count_backup':admin_hr_count_backup,
            'hallway_count_backup':hallway_count_backup,
            'it_room_storage_count_backup':it_room_storage_count_backup,
            'commentator_pantry_count_backup':commentator_pantry_count_backup,
            'ph_studio_count_backup':ph_studio_count_backup,
            'artist_dept_count_backup':artist_dept_count_backup,
            'seven_floor_count_backup':seven_floor_count_backup,
            'mini_pitx_count_backup': mini_pitx_count_backup,
            'plinko_count_backup':plinko_count_backup,
            #ready and backup count


            'las_vegas_ready_backup':las_vegas_ready_backup,
            'abu_dhabi_ready_backup':abu_dhabi_ready_backup,
            'suzuka_ready_backup': suzuka_ready_backup,
            'monaco_ready_backup':monaco_ready_backup,
            'c1_las_vegas_ready_backup': c1_las_vegas_ready_backup,
            'c2_abu_dhabi_ready_backup': c2_abu_dhabi_ready_backup,
            'c3_suzuka_ready_backup': c3_suzuka_ready_backup,
            'c4_monaco_ready_backup': c4_monaco_ready_backup,
            'it_room_ready_backup':it_room_ready_backup,
            'storage_1_ready_backup':storage_1_ready_backup,
            'admin_hr_ready_backup':admin_hr_ready_backup,
            'hallway_ready_backup':hallway_ready_backup,
            'it_room_storage_ready_backup':it_room_storage_ready_backup,
            'commentator_pantry_ready_backup':commentator_pantry_ready_backup,
            'ph_studio_ready_backup':ph_studio_ready_backup,
            'artist_dept_ready_backup':artist_dept_ready_backup,
            'seven_floor_ready_backup':seven_floor_ready_backup,
            'mini_pitx_ready_backup': mini_pitx_ready_backup,
            'plinko_ready_backup':plinko_ready_backup,

            'total_spare': total_spare,
            'overall_spare': overall_spare,
            'total_needed': total_needed,
            'total_needed_to_buy': total_needed_to_buy,

            #---------------------------------------------------------------------------------
            'requirement_las_vegas': self.request.session.get('requirement_las_vegas', 0),
            'propose_spare_las_vegas': self.request.session.get('propose_spare_las_vegas', 0),
            'las_vegas_total': las_vegas_total,
            #---------------------------------------------------------------------------------
            # ---------------------------------------------------------------------------------
            'requirement_abu_dhabi': self.request.session.get('requirement_abu_dhabi',0),
            'propose_spare_abu_dhabi': self.request.session.get('propose_spare_abu_dhabi',0),
            'abu_dhabi_total': abu_dhabi_total,
            # ---------------------------------------------------------------------------------
            'requirement_monaco': self.request.session.get('requirement_monaco',0),
            'propose_spare_monaco': self.request.session.get('propose_spare_monaco',0),
            'monaco_total': monaco_total,
            # ---------------------------------------------------------------------------------
            # ---------------------------------------------------------------------------------
            'requirement_suzuka': ('requirement_suzuka',0),
            'propose_spare_suzuka': ('propose_spare_suzuka',0),
            'suzuka_total': suzuka_total,
            # ---------------------------------------------------------------------------------
            'c1_requirement_las_vegas': self.request.session.get('c1_requirement_las_vegas', 0),
            'c1_propose_spare_las_vegas': self.request.session.get('c1_propose_spare_las_vegas', 0),
            'c1_las_vegas_total': c1_las_vegas_total,
            # ---------------------------------------------------------------------------------
            'c2_requirement_abu_dhabi': self.request.session.get('c2_requirement_abu_dhabi', 0),
            'c2_propose_spare_abu_dhabi': self.request.session.get('c2_propose_spare_abu_dhabi', 0),
            'c2_abu_dhabi_total': c2_abu_dhabi_total,
            # ---------------------------------------------------------------------------------
            'c4_requirement_monaco': self.request.session.get('c4_requirement_monaco', 0),
            'c4_propose_spare_monaco': self.request.session.get('c4_propose_spare_monaco', 0),
            'c4_monaco_total': c4_monaco_total,
            # ---------------------------------------------------------------------------------
            'c3_requirement_suzuka': ('c3_requirement_suzuka', 0),
            'c3_propose_spare_suzuka': ('c3_propose_spare_suzuka', 0),
            'c3_suzuka_total': c3_suzuka_total,
            # ---------------------------------------------------------------------------------

            'it_room_requirements': ('it_room_requirements', 0),
            'it_room_propose_spare': ('it_room_propose_spare', 0),
            'it_room_total': it_room_total,

            'it_room_storage_requirement': ('it_room_storage_requirement', 0),
            'it_room_storage_propose_spare': ('it_room_storage_propose_spare', 0),
            'it_room_storage_total': it_room_storage_total,

            # ---------------------------------------------------------------------------------
            'storage_1_requirement': ('storage_1_requirement', 0),
            'storage_1_propose_spare': ('storage_1_propose_spare', 0),
            'storage_1_total': storage_1_total,
            # ---------------------------------------------------------------------------------
            'admin_hr_requirement': ('admin_hr_requirement', 0),
            'admin_hr_propose_spare': ('admin_hr_propose_spare', 0),
            'admin_hr_total': admin_hr_total,
            # ---------------------------------------------------------------------------------
            'hallway_requirement': ('hallway_requirement', 0),
            'hallway_propose_spare': ('hallway_propose_spare', 0),
            'hallway_total': hallway_total,

            'ph_studio_requirement': ('ph_studio_requirement', 0),
            'ph_studio_propose_spare': ('ph_studio_propose_spare', 0),
            'ph_studio_total': ph_studio_total,

            'artist_dept_requirement': ('artist_dept_requirement', 0),
            'artist_dept_propose_spare': ('artist_dept_propose_spare', 0),
            'artist_dept_total': artist_dept_total,

            'seven_floor_requirement': ('seven_floor_requirement', 0),
            'seven_floor_propose_spare': ('seven_floor_propose_spare', 0),
            'seven_floor_total': seven_floor_total,

            'mini_pitx_requirement': ('mini_pitx_requirement', 0),
            'mini_pitx_propose_spare': ('mini_pitx_propose_spare', 0),
            'mini_pitx_total': mini_pitx_total,

            'plinko_requirement': ('plinko_requirement', 0),
            'plinko_propose_spare': ('plinko_propose_spare', 0),
            'plinko_total': plinko_total,


        })

        return context




    @transaction.atomic
    def post(self, request, pk):
        item = get_object_or_404(InventoryItem, pk=pk)

        def safe_int(value, default=0):
            try:
                return int(value)
            except ValueError:
                return default

        # Assign values with safe_int to handle empty strings or invalid values
        request.session['requirement_las_vegas'] = safe_int(request.POST.get('requirement_las_vegas', ''))
        request.session['propose_spare_las_vegas'] = safe_int(request.POST.get('propose_spare_las_vegas', ''))
        request.session['requirement_abu_dhabi'] = safe_int(request.POST.get('requirement_abu_dhabi', ''))
        request.session['propose_spare_abu_dhabi'] = safe_int(request.POST.get('propose_spare_abu_dhabi', ''))
        request.session['requirement_monaco'] = safe_int(request.POST.get('requirement_monaco', ''))
        request.session['propose_spare_monaco'] = safe_int(request.POST.get('propose_spare_monaco', ''))
        request.session['requirement_suzuka'] = safe_int(request.POST.get('requirement_suzuka', ''))
        request.session['propose_spare_suzuka'] = safe_int(request.POST.get('propose_spare_suzuka', ''))


        # Ensure the item has sufficient quantity
        if item.quantity <= 0:
            messages.error(request, f"No more quantity available for {item.name}.")
            return redirect('item-detail', pk=pk)

        # Decrease the inventory item's quantity
        item.quantity -= 1
        item.save()

        # Create a new SpecificItem
        SpecificItem.objects.create(
            inventory_item=item,
            added_by=request.user,
            quantity_added=1,  # Assuming 1 is always added
        )
        # Save user inputs in session

        # Add a success message and redirect
        return redirect('item-detail', pk=pk)

    def get_success_url(self):
        return reverse_lazy('item-detail', kwargs={'pk': self.object.pk})



class ItemDecrementView(View):
    def post(self, request, pk):
        # Get the item using the primary key (pk)
        item = get_object_or_404(InventoryItem, pk=pk)

        # Check if quantity is greater than 0 before decrementing
        if item.quantity > 0:
            item.quantity -= 1
            item.save()

            # Log the decrement in the transaction log
            TransactionLog.objects.create(
                user=request.user,
                action='decrement',
                description=f"Decremented quantity for item: {item.name}",
            )

            # Optionally, add to a specific dashboard
            SpecificItem.objects.create(
                inventory_item=item,
                added_by=request.user,
                quantity_added=1,
            )

         #   messages.success(request, f"Decremented quantity for {item.name} and added to the specific dashboard.")
     #   else:
        #    messages.error(request, f"{item.name} has no quantity left to decrement.")

        # Redirect back to the item detail page
        return redirect('item-detail', pk=item.pk)
class EditSpecificItemStatusView(UpdateView, TelegramNotificationMixin):
    model = SpecificItem
    form_class = SpecificItemStatusForm
    template_name = 'inventory/edit_specific_item_status.html'

    def get_success_url(self):
        return reverse_lazy('item-detail', kwargs={'pk': self.object.inventory_item.pk})

    def form_valid(self, form):
        specific_item = form.instance
        original_item = self.get_object()  # Get the original database record

        specific_item.last_modified_by = self.request.user

        # Capture old values
        old_status = original_item.status
        old_label = original_item.label
        old_room = original_item.room

        response = super().form_valid(form)

        # Capture new values after saving
        new_status = form.instance.status
        new_label = form.instance.label
        new_room = form.instance.room

        # Check if any values changed
        if old_status != new_status or old_label != new_label or old_room != new_room:
            # Log the change
            TransactionLog.objects.create(
                user=self.request.user,
                action='update',
                description=(
                    f"Updated {specific_item.inventory_item.name} - "
                    f"Status: {old_status} → {new_status}, "
                    f"Label: {old_label or 'None'} → {new_label or 'None'}, "
                    f"Room: {old_room or 'None'} → {new_room or 'None'}"
                ),
            )

            # Send Telegram alert
            self.send_telegram_alert(
                f"✅ Item: {specific_item.inventory_item.name} \n"
                f"✅ Tagging: {old_label or 'None'} → {new_label or 'None'}\n"
                f"✅ Room: {old_room or 'None'} → {new_room or 'None'}\n"
                f"✅ Status: {old_status} → {new_status}\n"
                f"✅ Updated by: {self.request.user.username}"
            )

        return response


class DeleteSpecificItemView(UserPassesTestMixin, View):
    def test_func(self):
        return self.request.user.is_superuser  # Only superusers can access this view

    def get(self, request, pk):
        specific_item = get_object_or_404(SpecificItem, pk=pk)
        return render(request, 'inventory/delete_specific_item_confirm.html', {'specific_item': specific_item})

    def post(self, request, pk):
        specific_item = get_object_or_404(SpecificItem, pk=pk)
        specific_item.delete()
        return HttpResponseRedirect(reverse('item-detail', args=[specific_item.inventory_item.pk]))


from django.http import JsonResponse
from django.views import View
from django.shortcuts import render
from django.contrib.auth.mixins import LoginRequiredMixin
import json

from .models import Lasvegas, Monaco, Suzuka, Abudhabi


class IncidentDashboardView(LoginRequiredMixin, View):
    template_name = "inventory/incident_dashboard.html"

    def get(self, request, *args, **kwargs):
        status_options = ['GOOD', 'NOT_GOOD', 'BROKEN']
        return render(request, self.template_name, {
            'statusOptions': status_options



        })

    def post(self, request, *args, **kwargs):
        data = json.loads(request.body)
        location = data.get('location')

        if location == 'lasvegas':
            model = Lasvegas
        elif location == 'monaco':
            model = Monaco
        elif location == 'suzuka':
            model = Suzuka
        elif location == 'abudhabi':
            model = Abudhabi
        elif location == 'philippines':
            model = Philippines
        elif location == 'plinko':
            model = Plinko
        else:
            return JsonResponse({'error': 'Invalid location'}, status=400)

        for entry in data.get('data', []):
            # ✅ Always create a new record, even if similar data exists
            model.objects.create(
                user=request.user,
                topview=entry.get('topview'),
                obs=entry.get('obs'),
                audio=entry.get('audio'),
                camera=entry.get('camera'),
                network_status=entry.get('network_status'),
                printer_status=entry.get('printer_status'),
                AI_Cam=entry.get('AI_Cam')


            )

        return JsonResponse({'status': 'success'}, status=200)


# views.py
from .models import Lasvegas, Monaco, Suzuka, Abudhabi

from datetime import datetime
from django.utils.timezone import now

def location_data_view(request):
    selected_date_str = request.GET.get('date')

    if selected_date_str:
        try:
            selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = now().date()
    else:
        selected_date = now().date()

    lasvegas_data = Lasvegas.objects.filter(timestamp__date=selected_date)
    monaco_data = Monaco.objects.filter(timestamp__date=selected_date)
    suzuka_data = Suzuka.objects.filter(timestamp__date=selected_date)
    abudhabi_data = Abudhabi.objects.filter(timestamp__date=selected_date)
    philippines_data = Philippines.objects.filter(timestamp__date=selected_date)
    plinko_data = Plinko.objects.filter(timestamp__date=selected_date)
    return render(request, 'inventory/location_data.html', {
        'lasvegas_data': lasvegas_data,
        'monaco_data': monaco_data,
        'suzuka_data': suzuka_data,
        'abudhabi_data': abudhabi_data,
        'philippines_data':philippines_data,
        'plinko_data': plinko_data,
        'selected_date': selected_date.strftime('%Y-%m-%d'),
        'today': now().date(),
    })

from .models import Lasvegas, Monaco, Suzuka, Abudhabi


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Lasvegas, Monaco, Suzuka, Abudhabi,Philippines,Plinko

@csrf_exempt
def clear_location_data(request):
    if request.method == "POST":
        location = request.POST.get('location')

        if location == 'lasvegas':
            Lasvegas.objects.all().delete()
        elif location == 'monaco':
            Monaco.objects.all().delete()
        elif location == 'suzuka':
            Suzuka.objects.all().delete()
        elif location == 'abudhabi':
            Abudhabi.objects.all().delete()
        elif location == 'philippines':
            Philippines.objects.all().delete()
        elif location == 'plinko':
            Plinko.objects.all().delete()
        else:
            return JsonResponse({'status': 'error', 'message': 'Invalid location'}, status=400)

        return JsonResponse({'status': 'success'})

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


# views.py
from django.views import View
from django.shortcuts import render, redirect
from django.contrib.auth.mixins import LoginRequiredMixin
from .models import Obstacle
from .forms import ObstacleForm

class ObstacleListView(LoginRequiredMixin, View):
    template_name = "inventory/obstacle-list.html"

    def get(self, request, *args, **kwargs):
        form = ObstacleForm()
        obstacles = Obstacle.objects.all()
        return render(request, self.template_name, {
            'form': form,
            'obstacles': obstacles
        })

    def post(self, request, *args, **kwargs):
        form = ObstacleForm(request.POST, request.FILES)
        if form.is_valid():
            obstacle = form.save(commit=False)
            obstacle.user = request.user
            obstacle.save()
            return redirect('obstacle-list')
        obstacles = Obstacle.objects.all()
        return render(request, self.template_name, {
            'form': form,
            'obstacles': obstacles
        })


from django.views.generic import DeleteView
from django.urls import reverse_lazy
from .models import Obstacle

class ObstacleDeleteView(LoginRequiredMixin, DeleteView):
    model = Obstacle
    success_url = reverse_lazy('obstacle-list')
    template_name = 'inventory/obstacle_confirm_delete.html'




# inventory/views.py
import uuid
import os
import requests

from datetime import date
from django.conf import settings
from django.shortcuts import render
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.template.loader import render_to_string
from django.http import HttpResponse
from .models import ChecklistSubmission

import os
import requests
from django.conf import settings
from django.template.loader import render_to_string
import imgkit

def send_telegram_notification(checklist_items, submission_details):
    """
    Generates a styled daily checklist image via Docker and sends it to Telegram.
    This function will fail loudly if image generation does not succeed.
    """
    image_filepath = None

    try:
        print("📡 Starting Telegram send...")
        bot_token = getattr(settings, 'TELEGRAM_BOT_TOKEN1')
        chat_id = getattr(settings, 'TELEGRAM_CHAT_ID1')

        # --- Generate Image ---
        image_filename = f"checklist_{submission_details['group_id']}.png"
        image_filepath = f"/tmp/{image_filename}"
        
        # ADD THIS LINE FOR DEBUGGING
        print("💡 USING CORRECT DOCKER-READY FUNCTION!") 

        print(f"🖼️ Generating image at {image_filepath} using Docker...")

        # Render HTML template (ensure this is correct)
        context = {'checklist_items': checklist_items, 'details': submission_details}
        inner_html = render_to_string('inventory/checklist_telegram_template.html', context)
        full_html = f"""
        <!DOCTYPE html><html lang="en"><head><meta charset="UTF-8"><style>
        body {{ background-color:#ffffff; font-family:Arial; padding:20px; }}
        h2 {{ text-align:center; color:#2f5496; font-weight:bold; }}
        table {{ width:100%; border-collapse:collapse; font-size:14px; }}
        th, td {{ border:1px solid #ccc; padding:8px; text-align:center; }}
        th {{ background-color:#2f5496; color:#fff; }}
        tr:nth-child(even) {{ background-color:#f9f9f9; }}
        </style></head><body>{inner_html}</body></html>
        """

        # Get the path to our Docker wrapper script from settings
        wkhtmltoimage_path = getattr(settings, 'WKHTMLTOIMAGE_CMD')
        if not os.path.exists(wkhtmltoimage_path):
             raise FileNotFoundError(f"FATAL: The wkhtmltoimage script was not found at {wkhtmltoimage_path}")
        
        config = imgkit.config(wkhtmltoimage=wkhtmltoimage_path)
        options = {'format': 'png', 'quality': '100', 'encoding': 'UTF-8', 'width': '1345', 'zoom': '1.5'}
        
        # This will now use the Docker script.
        imgkit.from_string(full_html, image_filepath, options=options, config=config)
        
        if not os.path.exists(image_filepath):
            raise IOError("Image generation command ran but the output file was not created.")

        print("✅ Image generated successfully via Docker.")

        # --- Send to Telegram ---
        url = f"https://api.telegram.org/bot{bot_token}/sendPhoto"
        with open(image_filepath, 'rb') as photo:
            caption = (
                f"✅ Daily Checklist submitted for {submission_details['studio']} "
                f"on {submission_details['date']} by {submission_details['user']}"
            )
            print("🚀 Sending image to Telegram...")
            response = requests.post(url, files={'photo': photo}, data={'chat_id': chat_id, 'caption': caption})

        print(f"📨 Telegram response: {response.status_code} - {response.text}")
        response.raise_for_status()
        print("✅ Telegram message sent successfully.")

    except Exception as e:
        print(f"❌ FATAL ERROR: Could not send Telegram image. Reason: {e}")

    finally:
        if image_filepath and os.path.exists(image_filepath):
            os.remove(image_filepath)
            print("🧹 Temporary image deleted.")


class Daily_Checklist_View(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        equipment_list = [
            "CONVEYOR", "RASPBERRY", "CAMERA", "STICKER", "START GATE",
            "GOAL GATE", "BOOSTER", "OBSTACLE", "SENSOR AND REFFLECTOR",
            "LCD CLOCK", "COMPRESSOR", "X AIR","CLEANING OF TRACK"
        ]
        context = {
            'page_title': 'Daily Operations Checklist',
            'equipments': equipment_list,
            'current_date': date.today().strftime("%Y-%m-%d")
        }
        return render(request, 'inventory/daily_checklist.html', context)

    def post(self, request, *args, **kwargs):
        studio = request.POST.get('studio')
        equipment_list = request.POST.getlist('equipment_name')
        submission_group_id = uuid.uuid4().hex[:10]

        # Collect entries
        items_to_create = []
        for i, equipment_name in enumerate(equipment_list, 1):
            items_to_create.append(ChecklistSubmission(
                submission_group_id=submission_group_id,
                studio=studio,
                equipment_name=equipment_name,
                checked_by=request.user,
                check_time=request.POST.get(f'time_{i}'),
                check_date=request.POST.get(f'date_{i}'),
                status=request.POST.get(f'status_{i}'),
                remarks=request.POST.get(f'remarks_{i}')
            ))

        # Save to DB
        ChecklistSubmission.objects.bulk_create(items_to_create)

        # ✅ Prepare data for Telegram (as readable dicts)
        checklist_items = []
        for item in items_to_create:
            checklist_items.append({
                'equipment_name': item.equipment_name,
                'user': item.checked_by.username if hasattr(item.checked_by, 'username') else str(item.checked_by),
                'status': item.status,
                'date': item.check_date if item.check_date else date.today().strftime("%Y-%m-%d"),
                'time': item.check_time if item.check_time else datetime.now().strftime("%I:%M %p"),
                'remarks': item.remarks or "—"
            })

        # ✅ Submission details
        submission_details = {
            'studio': studio,
            'date': date.today().strftime("%B %d, %Y"),
            'user': request.user.username,
            'group_id': submission_group_id
        }

        # ✅ Send to Telegram
        error_message = send_telegram_notification(checklist_items, submission_details)

        if error_message:
            return HttpResponse(f"❌ Failed to send Telegram image: {error_message}")
        else:
            return HttpResponse("✅ Checklist submitted and sent to Telegram successfully!")




from django.shortcuts import render
from django.contrib.auth.decorators import login_required

@login_required(login_url='/login/')
def looker_studio_dashboard(request):
    return render(request, "inventory/looker-studio.html")



from django.views import View
from django.shortcuts import render
from .models import SpecificItem

class BrokenItemsDashboardView(View):
    def get(self, request):
        broken_items = SpecificItem.objects.filter(status__iexact='Broken').select_related('inventory_item', 'inventory_item__user')
        return render(request, 'inventory/broken_items_dashboard.html', {
            'broken_items': broken_items,
            'query': request.GET.get('query', ''),
        })
